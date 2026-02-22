import pandas as pd
import json
import os
from datetime import datetime
from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from app.utils.cloudinary_service import upload_excel, download_from_cloudinary

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _apply_excel_style(path: str) -> None:
    """Apply bold header + yellow background (#FFFF00) to the first row of an Excel file."""
    try:
        wb = load_workbook(path)
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = yellow_fill
        wb.save(path)
    except Exception as e:
        print(f"⚠️  Excel styling failed for {path}: {e}")


def normalize_name(name) -> str | None:
    """Normalize a client name: strip whitespace, uppercase. Returns None if null."""
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return None
    return str(name).strip().upper()


# ---------------------------------------------------------------------------
# Phase-1: Two-file upload with email validation
# ---------------------------------------------------------------------------

def process_master_with_email_validation(
    master_path: str,
    email_path: str,
    batch_id: str,
    output_dir: str,
) -> dict:
    """
    Phase-1 processing:
    1. Read master file, extract unique client names from 'Order id' column.
    2. Read email mapping file (Client_Name, Client_Email).
    3. Validate every master client has a mapped email.
    4. On success: save client_email_map.json and master_with_clients.xlsx.
    """

    # -- 1. Read master file --------------------------------------------------
    try:
        master_df = pd.read_excel(master_path)
    except Exception as e:
        return {"success": False, "message": f"Cannot read master file: {e}"}

    # -- 2. Validate 'Order id' column ----------------------------------------
    if "Order id" not in master_df.columns:
        return {
            "success": False,
            "message": "Master file is missing required column: 'Order id'",
        }

    # -- 3. Create normalized Client_Name column -------------------------------
    master_df["Client_Name"] = master_df["Order id"].apply(normalize_name)

    # -- 4. Unique client names (drop None/blank) ------------------------------
    master_clients = set(
        name for name in master_df["Client_Name"].dropna().unique() if name != ""
    )

    if not master_clients:
        return {"success": False, "message": "Master file contains no client names in 'Order id' column."}

    # -- 5. Read email mapping file -------------------------------------------
    try:
        email_df = pd.read_excel(email_path)
    except Exception as e:
        return {"success": False, "message": f"Cannot read email mapping file: {e}"}

    # -- 6. Normalize column names (flexible: handles spaces, mixed case, underscores)
    # e.g. "Client Name", "client name", "CLIENT NAME", "Client_Name" → "client_name"
    email_df.columns = [col.strip().lower().replace(" ", "_") for col in email_df.columns]

    required_email_cols = ["client_name", "client_email"]
    for col in required_email_cols:
        if col not in email_df.columns:
            return {
                "success": False,
                "message": "Email mapping file must contain columns: Client Name, Client Email",
            }

    # Rename to standard internal format
    email_df.rename(columns={"client_name": "Client_Name", "client_email": "Client_Email"}, inplace=True)

    # -- 7. Normalize Client_Name in email file --------------------------------
    email_df["Client_Name"] = email_df["Client_Name"].apply(normalize_name)

    # -- 8. Build mapping dict: {Client_Name -> Client_Email} ------------------
    client_email_map: dict[str, str] = {}
    for _, row in email_df.iterrows():
        name = row["Client_Name"]
        email = str(row["Client_Email"]).strip() if pd.notnull(row["Client_Email"]) else ""
        if name and email:
            client_email_map[name] = email

    email_clients = set(client_email_map.keys())

    # -- 9. Compare sets -------------------------------------------------------
    missing_clients = master_clients - email_clients

    # -- 10. Return error if any client is unmapped ----------------------------
    if missing_clients:
        return {
            "success": False,
            "message": f"Emails missing for clients: {sorted(missing_clients)}",
        }

    # -- 11. Save outputs ------------------------------------------------------
    os.makedirs(output_dir, exist_ok=True)

    # Save client_email_map.json
    map_path = os.path.join(output_dir, "client_email_map.json")
    with open(map_path, "w") as f:
        json.dump(client_email_map, f, indent=4)

    # Save master_with_clients.xlsx
    enriched_path = os.path.join(output_dir, "master_with_clients.xlsx")
    master_df.to_excel(enriched_path, index=False)

    return {"success": True}


# ---------------------------------------------------------------------------
# Phase-2: Generate Mother File + Split Client Files
# ---------------------------------------------------------------------------

# Fixed mapping: output column name → exact master column name (strip only).
# "DATE" uses Pickup Date only — Manifest Date is intentionally excluded to avoid duplicates.
COLUMN_MAPPING = {
    "DATE":              "Pickup Date",
    "C/NO":              "LRN",
    "C/NOR":             "Order id",
    "C/NEE":             "Consignee name",
    "DEST":              "Destination City",
    "PIN CODE":          "Pin code",
    "INVOICE NO":        "Invoice Number",
    "TYPE":              "Transaction Type",
    "QUTY":              "No of boxes",
    "STATUS":            "Current Status",
    "D DATE":            "Delivered Date",
    "REMARKS":           "Remarks",
    "Expected delivery": "Expected Date",
}

# Final column order that every client/mother file must conform to.
FINAL_COLUMNS = [
    "DATE",
    "C/NO",
    "C/NOR",
    "C/NEE",
    "DEST",
    "PIN CODE",
    "INVOICE NO",
    "TYPE",
    "QUTY",
    "STATUS",
    "D DATE",
    "REMARKS",
    "Expected delivery",
]


def _clean_mis_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Standardize a MIS dataframe before saving to Excel:
    1. Fill NaN / 'nan' / 'None' / '#' with empty string.
    2. Format date columns as DD-MM-YYYY.
    3. Strip trailing '.0' from PIN CODE and QUTY.
    4. Strip leading/trailing spaces from all string columns.
    """
    # -- Missing values -------------------------------------------------------
    df = df.fillna("")
    for bad in ("nan", "None", "#", "NaT"):
        df.replace(bad, "", inplace=True)

    # -- Date formatting ------------------------------------------------------
    for col in ["DATE", "D DATE", "Expected delivery"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d-%m-%Y")
            df[col] = df[col].fillna("")          # NaT → ""
            df[col].replace("NaT", "", inplace=True)

    # -- Numeric cleanup (remove spurious '.0') --------------------------------
    for col in ["PIN CODE", "QUTY"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace("\.0$", "", regex=True)
            df[col] = df[col].replace("nan", "").replace("None", "")

    # -- Strip whitespace from all string columns -----------------------------
    df = df.apply(lambda c: c.str.strip() if c.dtype == object else c)

    return df


def _safe_filename(client_name: str) -> str:
    """Convert a client name to a safe filename.
    Uppercases, replaces spaces with underscores, removes / and \\.
    """
    return (
        client_name.upper()
        .replace(" ", "_")
        .replace("/", "")
        .replace("\\", "")
    )


def generate_client_mis_files(batch_folder: str, batch_id: str = None) -> dict:
    """
    Phase-2 — Generate Mother File + Split Client Files:
    1. Load master_with_clients.xlsx.
    2. Strip-normalize master column names.
    3. Build mother_df using fixed COLUMN_MAPPING (mother_col → master_col).
    4. Add Client_Name from 'Client' or 'Order id' column.
    5. Save mother.xlsx to processed/.
    6. Split by Client_Name and save one Excel per client to client_files/.

    Returns:
        {"success": True, "total_rows": N, "total_clients": N, "files_created": [...]}
     or {"success": False, "message": "..."}
    """
    master_path = os.path.join(batch_folder, "master_with_clients.xlsx")

    # -- 1. Load master_with_clients.xlsx ------------------------------------
    if not os.path.exists(master_path):
        return {
            "success": False,
            "message": "master_with_clients.xlsx not found. Ensure Phase-1 completed successfully.",
        }

    try:
        df = pd.read_excel(master_path)
    except Exception as e:
        return {"success": False, "message": f"Cannot read master_with_clients.xlsx: {e}"}

    # -- 2. Strip-normalize master column names (exact match, strip only) ----
    df.columns = [str(col).strip() for col in df.columns]

    # -- 3. Build mother_df using fixed COLUMN_MAPPING -----------------------
    # COLUMN_MAPPING keys are already the final output column names.
    mother_df = pd.DataFrame(index=df.index)

    for out_col, master_col in COLUMN_MAPPING.items():
        if master_col in df.columns:
            mother_df[out_col] = df[master_col]
        else:
            mother_df[out_col] = ""   # column truly missing — keep empty

    # -- 4. Add Client_Name from 'Order id' column (only) -------------------
    if "Order id" not in df.columns:
        return {
            "success": False,
            "message": "Master file is missing required column: 'Order id'",
        }

    mother_df["Client_Name"] = df["Order id"].astype(str).str.strip()

    # -- 4b. Enforce FINAL_COLUMNS order (all guaranteed present from mapping) ---
    mother_df = mother_df[FINAL_COLUMNS + ["Client_Name"]]

    # -- 4c. Clean missing values + format --------------------------------
    mother_df = _clean_mis_df(mother_df)

    # -- 5. Save mother.xlsx temporarily, upload to Cloudinary, delete local --
    if batch_id is None:
        batch_id = os.path.basename(batch_folder)

    tmp_dir = os.path.join("app", "storage", "tmp", batch_id)
    os.makedirs(tmp_dir, exist_ok=True)

    mother_tmp = os.path.join(tmp_dir, "mother.xlsx")
    mother_df.to_excel(mother_tmp, index=False)
    _apply_excel_style(mother_tmp)

    try:
        mother_cloud = upload_excel(mother_tmp, batch_id, "mother", "mother")
    except Exception as e:
        return {"success": False, "message": f"Cloudinary upload failed for mother.xlsx: {e}"}
    finally:
        if os.path.exists(mother_tmp):
            os.remove(mother_tmp)

    total_rows = len(mother_df)

    # -- 6. Split client-wise, upload each to Cloudinary, no local retention --
    groups = mother_df.groupby("Client_Name", dropna=True)

    total_clients = 0
    client_files_cloud: dict = {}

    for client, group_df in groups:
        # Skip blank or NaN string values
        if not client or client.lower() == "nan":
            continue

        safe_name = client.replace(" ", "_").replace("/", "_")
        tmp_path = os.path.join(tmp_dir, f"{safe_name}.xlsx")

        client_df = group_df.drop(columns=["Client_Name"])
        # Enforce final column order, format and clean
        client_df = client_df[FINAL_COLUMNS]
        client_df = _clean_mis_df(client_df)
        client_df.to_excel(tmp_path, index=False)
        _apply_excel_style(tmp_path)

        try:
            cloud_info = upload_excel(tmp_path, batch_id, "client_files", safe_name)
            client_files_cloud[safe_name] = cloud_info
        except Exception as e:
            print(f"⚠️  Cloudinary upload failed for {safe_name}: {e}")
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

        total_clients += 1

    # Clean up tmp batch dir if now empty
    try:
        if os.path.exists(tmp_dir) and not os.listdir(tmp_dir):
            os.rmdir(tmp_dir)
    except Exception:
        pass

    if total_clients == 0:
        return {
            "success": False,
            "message": "No client files were generated (all client names were blank).",
        }

    return {
        "success": True,
        "total_rows": total_rows,
        "master_rows": len(df),
        "total_clients": total_clients,
        "mother_columns": list(mother_df.columns),
        "mother_url": mother_cloud.get("url"),
        "mother_public_id": mother_cloud.get("public_id"),
        "client_files": client_files_cloud,
    }


# ---------------------------------------------------------------------------
# Phase-3: Send MIS emails to each client with their Excel file attached
# ---------------------------------------------------------------------------

def send_client_mis_emails(batch_folder: str, clients: list = None, limit: int = None, file_type: str = "generated") -> dict:
    """
    Phase-3 processing:
    1. Load client_email_map.json from the batch folder.
    2. Locate Excel files in client_files/ subfolder.
    3. Filter to only requested clients (if clients list is provided).
    4. For each client file, send an email via AWS SES with the file attached.
    5. Log all results to batch_folder/email_logs.txt.
    6. Respect optional limit on number of emails sent.

    Returns:
        {"success": True,  "total_sent": N, "failed": M, "errors": [...]}
     or {"success": False, "message": "..."}
    """
    import boto3
    import re as _re
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    from app.config import settings

    # -- 1. Load email mapping ------------------------------------------------
    mapping_path = os.path.join(batch_folder, "client_email_map.json")
    if not os.path.exists(mapping_path):
        return {
            "success": False,
            "message": "client_email_map.json not found. Ensure Phase-1 completed successfully.",
        }

    try:
        with open(mapping_path, "r") as f:
            email_map: dict = json.load(f)   # {"CLIENT NAME": "email@x.com"}
    except Exception as e:
        return {"success": False, "message": f"Cannot read client_email_map.json: {e}"}

    # -- 2. Load Cloudinary URLs from meta.json --------------------------------
    meta_path = os.path.join(batch_folder, "meta.json")
    if not os.path.exists(meta_path):
        return {
            "success": False,
            "message": "meta.json not found. Ensure Phase-2 completed successfully.",
        }

    try:
        with open(meta_path, "r") as _mf:
            meta = json.load(_mf)
    except Exception as e:
        return {"success": False, "message": f"Cannot read meta.json: {e}"}

    cloud_generated: dict = meta.get("client_files", {})
    cloud_custom: dict = meta.get("custom_files", {})

    if not cloud_generated:
        return {"success": False, "message": "No client files found in meta.json. Ensure Phase-2 completed successfully."}

    all_safe_names = list(cloud_generated.keys())

    # -- 3a. Filter to selected clients only (if a list was provided) ---------
    if clients:
        requested_safe  = {c.upper().strip().replace(" ", "_") for c in clients}
        requested_space = {c.upper().strip() for c in clients}
        all_safe_names = [
            sn for sn in all_safe_names
            if sn.upper() in requested_safe or sn.replace("_", " ").upper() in requested_space
        ]
        if not all_safe_names:
            return {"success": False, "message": "None of the selected clients have MIS files."}

    # Prepare tmp directory for downloads
    tmp_dir = os.path.join("app", "storage", "tmp", os.path.basename(batch_folder))
    os.makedirs(tmp_dir, exist_ok=True)

    # -- 3. Initialise SES client ---------------------------------------------
    sender_email = settings.SES_SENDER_EMAIL
    sender_name  = settings.SES_SENDER_NAME
    sender       = f"{sender_name} <{sender_email}>"

    try:
        ses = boto3.client(
            "ses",
            region_name=settings.AWS_REGION,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        )
    except Exception as e:
        return {"success": False, "message": f"Failed to create SES client: {e}"}

    # -- 4. Prepare log file --------------------------------------------------
    log_path = os.path.join(batch_folder, "email_logs.txt")

    def _log(client_name: str, recipient: str, status: str, error: str = None):
        timestamp = datetime.now().isoformat()
        entry = f"{timestamp} | {client_name} | {recipient} | {status}"
        if error:
            entry += f" | {error}"
        with open(log_path, "a") as lf:
            lf.write(entry + "\n")

    # -- 5. Send emails -------------------------------------------------------
    today_date = datetime.now().strftime("%d %b %Y")
    total_sent  = 0
    failed      = 0
    errors      = []
    sent_clients   = []   # list of {safe_name, client_name, email}
    failed_clients = []   # list of {safe_name, client_name, email, reason}

    for safe_name in all_safe_names:
        # Respect limit
        if limit is not None and total_sent >= limit:
            break

        # Derive display client name: PERKINS_INDIA → PERKINS INDIA
        client_name = safe_name.replace("_", " ").strip()

        # Look up email — try exact match first, then uppercase
        recipient = email_map.get(client_name) or email_map.get(client_name.upper())
        if not recipient:
            print(f"⚠️  No email found for client '{client_name}' — skipping.")
            continue

        # Determine which Cloudinary URL to use
        if file_type == "custom" and safe_name in cloud_custom:
            cloud_url = cloud_custom[safe_name]["url"]
        elif safe_name in cloud_generated:
            cloud_url = cloud_generated[safe_name]["url"]
        else:
            print(f"⚠️  No Cloudinary URL for '{client_name}' — skipping.")
            continue

        # Download from Cloudinary to temp file
        filename = f"{safe_name}.xlsx"
        file_path = os.path.join(tmp_dir, filename)
        try:
            download_from_cloudinary(cloud_url, file_path)
        except Exception as e:
            failed += 1
            error_msg = f"Cloudinary download failed: {e}"
            errors.append({"client": client_name, "email": recipient, "reason": error_msg})
            _log(client_name, recipient, "Failed", error_msg)
            print(f"❌ {client_name}: {error_msg}")
            continue

        # Build MIME message
        subject = f"{client_name} – MIS – {today_date}"
        body = (
            f"Dear Sir,\n\n"
            f"Kindly find attached MIS for last 30 days.\n\n"
            f"Regards,\nKiirusxpress Team"
        )

        msg = MIMEMultipart()
        msg["Subject"] = subject
        msg["From"]    = sender
        msg["To"]      = recipient
        msg.attach(MIMEText(body, "plain"))

        with open(file_path, "rb") as xf:
            part = MIMEApplication(xf.read())
            part.add_header("Content-Disposition", "attachment", filename=filename)
            msg.attach(part)

        # Send via SES raw email
        try:
            ses.send_raw_email(
                Source=sender,
                Destinations=[recipient],
                RawMessage={"Data": msg.as_string()},
            )
            total_sent += 1
            sent_clients.append({"safe_name": safe_name, "client_name": client_name, "email": recipient})
            _log(client_name, recipient, "Sent")
            print(f"✅ Sent MIS to {client_name} <{recipient}>")
        except Exception as e:
            failed += 1
            error_msg = str(e)
            errors.append({"client": client_name, "email": recipient, "reason": error_msg})
            failed_clients.append({"safe_name": safe_name, "client_name": client_name, "email": recipient, "reason": error_msg})
            _log(client_name, recipient, "Failed", error_msg)
            print(f"❌ Failed to send to {client_name}: {error_msg}")
        finally:
            # Delete temp download immediately after sending
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass

    # Clean up tmp dir for this batch
    try:
        if os.path.exists(tmp_dir) and not os.listdir(tmp_dir):
            os.rmdir(tmp_dir)
    except Exception:
        pass

    # -- 6. Return summary ----------------------------------------------------
    return {
        "success": True,
        "total_sent": total_sent,
        "failed": failed,
        "errors": errors,
        "sent_clients": sent_clients,
        "failed_clients": failed_clients,
    }


# ---------------------------------------------------------------------------
# Legacy: original single-file processing (kept for backward compatibility)
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = [
    "Customer_Name",
    "Customer_Email",
    "Parcel_Count",
    "Dispatch_Date",
    "Total_Weight",
    "Payment_Status"
]

COLUMNS_TO_REMOVE = [
    "Customer_ID",
    "Freight_Amount",
    "Sales_Engineer",
    "Internal_Remarks",
    "Customer_Phone"
]

def process_master_file(file_path: str, output_dir: str, batch_id: str):
    try:
        # 1. Read Excel
        df = pd.read_excel(file_path)
        
        # 2. Validate columns
        missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing_columns)}")

        # 3. Remove unwanted columns if they exist
        drop_cols = [col for col in COLUMNS_TO_REMOVE if col in df.columns]
        df = df.drop(columns=drop_cols)
        
        # 3.5 Save Mother File (Requirement)
        mother_file_path = os.path.join(output_dir, "mother.xlsx")
        df.to_excel(mother_file_path, index=False)

        # 4. Process each row individually (NO GROUPING)
        rows = []
        
        # Ensure Dispatch_Date is datetime for proper formatting
        if 'Dispatch_Date' in df.columns:
            df['Dispatch_Date'] = pd.to_datetime(df['Dispatch_Date'], errors='coerce')

        # Iterate through each row in the dataframe
        for idx, row in df.iterrows():
            # Extract data from each row
            customer_name = row.get("Customer_Name", "Unknown")
            customer_email = str(row.get("Customer_Email", "")).strip()
            parcel_count = int(row.get("Parcel_Count", 0))
            total_weight = float(row.get("Total_Weight", 0.0))
            payment_status = row.get("Payment_Status", "")
            
            # Format dispatch date
            dispatch_date = None
            if "Dispatch_Date" in row and pd.notnull(row["Dispatch_Date"]):
                dispatch_date = row["Dispatch_Date"].strftime("%Y-%m-%d")
            
            # Create row record
            row_record = {
                "row_id": idx + 1,  # 1-based row ID
                "customer_name": customer_name,
                "customer_email": customer_email,
                "parcel_count": parcel_count,
                "total_weight": total_weight,
                "dispatch_date": dispatch_date,
                "payment_status": payment_status,
                "status": "NotSent"
            }
            rows.append(row_record)

        # 5. Save Row Data
        # summary_filename = f"{timestamp}.json" # Old way
        summary_path = os.path.join(output_dir, "summary.json")
        
        # Ensure directory exists just in case (though upload handles it)
        os.makedirs(output_dir, exist_ok=True)
        
        batch_data = {
            "batch_id": batch_id,
            "created_at": datetime.now().isoformat(),
            "rows": rows,  # Changed from "customers" to "rows"
            "total_rows": len(rows)
        }
        
        with open(summary_path, "w") as f:
            json.dump(batch_data, f, indent=4)

        # Create email_log.json for tracking
        email_log_path = os.path.join(output_dir, "email_log.json")
        email_log = {
            "batch_id": batch_id,
            "created_at": datetime.now().isoformat(),
            "emails": []
        }
        with open(email_log_path, "w") as f:
            json.dump(email_log, f, indent=4)

        return {
            "message": "Master processed successfully",
            "batch_id": batch_id,
            "total_rows": len(rows)
        }

    except Exception as e:
        # Re-raise HTTP exceptions, wrap others
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))
