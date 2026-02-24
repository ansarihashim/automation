"""
MIS (Management Information System) generation service.

Generates per-client Excel MIS files sourced entirely from the cumulative
'shipments' MongoDB collection — not from the raw uploaded master file.

Each file:
  • Has one sheet per calendar month (Jan-2026, Feb-2026, …)
  • Applies bold + yellow header row with thin cell borders
  • Is uploaded to Cloudinary under  kiirus/mis/<SAFE_NAME>
  • Local /tmp copy is deleted after upload

Public API
----------
    generate_client_mis(client_name)       → Cloudinary URL (str)
    generate_mis_for_all_clients()         → {client_name: {"url": ..., "public_id": ...}}
"""

from __future__ import annotations

import math
import os
import tempfile
from datetime import datetime, timezone
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.utils.cloudinary_service import upload_mis_file

# ---------------------------------------------------------------------------
# Column configuration
# ---------------------------------------------------------------------------

# Maps  output-column-name → MongoDB field name
_COLUMN_MAP: dict[str, str] = {
    "BOOKING DATE":      "manifest_date",
    "LRN":               "lrn",
    "CONSIGNOR NAME":    "order_id",       # Order reference (distinct from LRN)
    "CONSIGNEE NAME":    "consignee_name",
    "ORIGIN":            "origin",
    "DESTINATION":       "destination",
    "PIN CODE":          "pin_code",
    "INVOICE NO":        "invoice_number",
    "NO OF BOXES":       "no_of_boxes",   # post-processed (–1, clamp ≥ 0)
    "STATUS":            "status",
    "DATE OF DELIVERY":  "delivered_date",
    "REMARKS":           "remarks",
    "EXPECTED DELIVERY": "expected_date",
}

FINAL_COLUMNS: list[str] = [
    "BOOKING DATE",
    "LRN",
    "CONSIGNOR NAME",
    "CONSIGNEE NAME",
    "ORIGIN",
    "DESTINATION",
    "PIN CODE",
    "INVOICE NO",
    "NO OF BOXES",
    "STATUS",
    "DATE OF DELIVERY",
    "REMARKS",
    "EXPECTED DELIVERY",
]

# Date columns that must be formatted as DD-MM-YYYY strings in the output
_DATE_COLS = {"BOOKING DATE", "DATE OF DELIVERY", "EXPECTED DELIVERY"}

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _safe_sheet_name(period: pd.Period) -> str:
    """Convert pandas Period to readable sheet name, e.g. Jan-2026."""
    return period.strftime("%b-%Y")


def _format_df(raw: list[dict]) -> pd.DataFrame:
    """
    Convert a list of MongoDB shipment documents into the final MIS DataFrame.

    Steps
    -----
    1. Build a DataFrame from raw documents.
    2. Map MongoDB fields → output column names.
    3. Compute NO OF BOXES = max(0, no_of_boxes - 1).
    4. Format date columns as DD-MM-YYYY strings (blank for NaT / None).
    5. Clean NaN → empty string for remaining string columns.
    6. Return only FINAL_COLUMNS in declared order.
    """
    if not raw:
        return pd.DataFrame(columns=FINAL_COLUMNS)

    src = pd.DataFrame(raw)

    out = pd.DataFrame(index=src.index)

    for col_out, col_src in _COLUMN_MAP.items():
        if col_src in src.columns:
            out[col_out] = src[col_src]
        else:
            out[col_out] = ""

    # ── NO OF BOXES: subtract 1, clamp to 0 ─────────────────────────────────
    out["NO OF BOXES"] = (
        pd.to_numeric(out["NO OF BOXES"], errors="coerce")
        .fillna(0)
        .astype(int)
        .subtract(1)
        .clip(lower=0)
    )

    # ── Date columns → DD-MM-YYYY strings ────────────────────────────────────
    for date_col in _DATE_COLS:
        out[date_col] = (
            pd.to_datetime(out[date_col], errors="coerce")
            .dt.strftime("%d-%m-%Y")
            .fillna("")
        )
        out[date_col] = out[date_col].replace("NaT", "")

    # ── Remaining string columns: NaN → "" ───────────────────────────────────
    for col in FINAL_COLUMNS:
        if col not in _DATE_COLS and col != "NO OF BOXES":
            out[col] = out[col].fillna("").astype(str).str.strip()
            out[col] = out[col].replace("nan", "").replace("None", "")

    return out[FINAL_COLUMNS]


def _apply_formatting(ws) -> None:
    """
    Apply to an openpyxl worksheet:
      • Row 1: bold font + yellow (#FFFF00) fill
      • All cells: thin border on all 4 sides
    """
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    bold_font   = Font(bold=True)
    thin_side   = Side(border_style="thin", color="000000")
    thin_border = Border(
        left=thin_side, right=thin_side,
        top=thin_side,  bottom=thin_side,
    )

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = thin_border
            if row_idx == 1:
                cell.font = bold_font
                cell.fill = yellow_fill

    # Auto-fit column widths (best-effort)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        # Clamp width between 10 and 40 characters
        adjusted_width = min(max(max_len + 2, 10), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


def _build_excel(df: pd.DataFrame, tmp_path: str) -> None:
    """
    Write *df* to *tmp_path* as a multi-sheet Excel workbook.

    Sheets are split by BOOKING DATE month. Rows where BOOKING DATE cannot be
    parsed (blank / invalid) are collected in an "Unknown" sheet so no data
    is silently dropped.
    """
    # Parse BOOKING DATE back to period for grouping (it's already a string)
    df = df.copy()
    df["_period"] = pd.to_datetime(df["BOOKING DATE"], format="%d-%m-%Y", errors="coerce").dt.to_period("M")

    has_unknown = df["_period"].isna().any()

    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
        # ── Monthly sheets (sorted chronologically) ──────────────────────────
        periods = sorted(df["_period"].dropna().unique())

        if not periods and not has_unknown:
            # Edge case: completely empty BOOKING DATE → write everything to one sheet
            _write_sheet(writer, df.drop(columns=["_period"]), "All")
        else:
            for period in periods:
                subset = df[df["_period"] == period].drop(columns=["_period"])
                _write_sheet(writer, subset, _safe_sheet_name(period))

            if has_unknown:
                subset = df[df["_period"].isna()].drop(columns=["_period"])
                _write_sheet(writer, subset, "Unknown")

    # Post-process: apply openpyxl formatting (borders + header style)
    wb = load_workbook(tmp_path)
    for ws in wb.worksheets:
        _apply_formatting(ws)
    wb.save(tmp_path)


def _write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    """Write one DataFrame to one sheet without the pandas index."""
    # Trim sheet name to Excel's 31-char limit
    safe = sheet_name[:31]
    df.to_excel(writer, sheet_name=safe, index=False)


def _safe_name(client_name: str) -> str:
    """Convert 'PERKINS INDIA' → 'PERKINS_INDIA'."""
    return client_name.strip().upper().replace(" ", "_").replace("/", "_")


# ---------------------------------------------------------------------------
# Public service functions
# ---------------------------------------------------------------------------

async def generate_client_mis(client_name: str) -> dict:
    """
    Generate a per-client MIS Excel file from MongoDB and upload to Cloudinary.

    Parameters
    ----------
    client_name : str
        Exact value as stored in shipments.client_name (normalised uppercase).

    Returns
    -------
    dict with keys:
        url        – Cloudinary secure_url
        public_id  – Cloudinary public_id

    Raises
    ------
    ValueError  if no shipments exist for this client.
    RuntimeError on Cloudinary upload failure.
    """
    db = get_db()

    # ── Fetch all shipments for this client ──────────────────────────────────
    cursor = db["shipments"].find(
        {"client_name": client_name},
        {"_id": 0},  # exclude MongoDB _id
    )
    raw_docs: list[dict] = await cursor.to_list(length=None)

    if not raw_docs:
        raise ValueError(f"No shipments found for client: {client_name!r}")

    print(f"  📄 {client_name}: {len(raw_docs)} shipment records fetched")

    # ── Build MIS DataFrame ──────────────────────────────────────────────────
    df = _format_df(raw_docs)

    # ── Write Excel to /tmp ───────────────────────────────────────────────────
    safe = _safe_name(client_name)
    filename = f"{safe}_MIS.xlsx"
    tmp_path = os.path.join(tempfile.gettempdir(), filename)

    try:
        _build_excel(df, tmp_path)
    except Exception as exc:
        raise RuntimeError(f"Excel generation failed for {client_name!r}: {exc}") from exc

    # ── Upload to Cloudinary & clean up ─────────────────────────────────────
    try:
        cloud_info = upload_mis_file(tmp_path, client_name)
    finally:
        # Always delete the local temp file, even if upload fails
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass

    print(f"  ☁️  {client_name}: uploaded → {cloud_info['url']}")
    return cloud_info


async def generate_mis_for_all_clients() -> dict[str, dict]:
    """
    Generate MIS files for every distinct client_name in the 'shipments'
    collection and upload each to Cloudinary.

    Returns
    -------
    dict mapping  client_name → {"url": ..., "public_id": ...}

    Per-client failures are caught and logged; they do NOT abort the loop so
    a single bad client cannot block the entire batch.
    """
    db = get_db()

    client_names: list[str] = await db["shipments"].distinct("client_name")
    # Filter out None / blank values
    client_names = [c for c in client_names if c and str(c).strip()]

    if not client_names:
        print("⚠️  generate_mis_for_all_clients: no clients found in shipments collection.")
        return {}

    print(f"\n📊 Generating MIS for {len(client_names)} client(s)…")

    results: dict[str, dict] = {}

    for client_name in sorted(client_names):
        try:
            cloud_info = await generate_client_mis(client_name)
            results[client_name] = cloud_info
        except Exception as exc:
            print(f"  ❌ MIS generation failed for {client_name!r}: {exc}")
            results[client_name] = {"url": None, "public_id": None, "error": str(exc)}

    ok_count   = sum(1 for v in results.values() if v.get("url"))
    fail_count = len(results) - ok_count
    print(
        f"✅ MIS generation complete — "
        f"Success: {ok_count} | Failed: {fail_count} | Total: {len(results)}\n"
    )

    return results
